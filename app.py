#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
仓库管理系统 Web API
"""

from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_from_directory
from flask_cors import CORS
from database import WarehouseDB
from datetime import datetime
import secrets
import os

app = Flask(__name__, static_folder='static', static_url_path='')
app.secret_key = secrets.token_hex(16)

# 配置 CORS 支持开发环境
CORS(app, supports_credentials=True, origins=['http://localhost:3000'])

db = WarehouseDB()

# 登录验证装饰器
def login_required(f):
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            return jsonify({'success': False, 'message': '请先登录', 'redirect': '/login'}), 401
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

# 管理员权限验证装饰器
def admin_required(f):
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            return jsonify({'success': False, 'message': '请先登录'}), 401
        if session['user']['role'] not in ['ADMIN', 'MANAGER']:
            return jsonify({'success': False, 'message': '权限不足'}), 403
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

@app.route('/')
def index():
    """主页 - 返回 Vue 应用"""
    return send_from_directory('static', 'index.html')

@app.route('/login')
def login_page():
    """登录页面 - 返回 Vue 应用"""
    return send_from_directory('static', 'index.html')

# 处理 Vue Router 的所有路由
@app.route('/<path:path>')
def catch_all(path):
    """捕获所有路由，返回 Vue 应用（用于 Vue Router 的 history 模式）"""
    # 如果是 API 请求，返回 404
    if path.startswith('api/'):
        return jsonify({'success': False, 'message': 'API endpoint not found'}), 404
    # 如果请求的是静态文件，尝试返回
    if '.' in path:
        try:
            return send_from_directory('static', path)
        except:
            pass
    # 其他所有请求返回 index.html（Vue Router 会处理）
    return send_from_directory('static', 'index.html')

@app.route('/api/login', methods=['POST'])
def login():
    """用户登录"""
    data = request.json
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    
    if not username or not password:
        return jsonify({'success': False, 'message': '用户名和密码不能为空'})
    
    success, user = db.login(username, password)
    
    if success:
        session['user'] = user
        return jsonify({'success': True, 'message': '登录成功', 'user': user})
    else:
        return jsonify({'success': False, 'message': '用户名或密码错误'})

@app.route('/api/logout', methods=['POST'])
def logout():
    """用户登出"""
    session.pop('user', None)
    return jsonify({'success': True, 'message': '已退出登录'})

@app.route('/api/current-user', methods=['GET'])
@login_required
def get_current_user():
    """获取当前登录用户"""
    return jsonify({'success': True, 'user': session['user']})

@app.route('/api/users', methods=['GET'])
@admin_required
def get_users():
    """获取所有用户"""
    role = request.args.get('role', None)
    users = db.get_all_users(role=role)
    return jsonify({
        'success': True,
        'data': [{
            'user_id': u[0],
            'username': u[1],
            'real_name': u[2],
            'role': u[3],
            'created_at': u[4]
        } for u in users]
    })

@app.route('/api/users', methods=['POST'])
@admin_required
def add_user():
    """添加用户"""
    data = request.json
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    real_name = data.get('real_name', '').strip()
    role = data.get('role', '').strip()
    
    if not username or not password or not real_name or not role:
        return jsonify({'success': False, 'message': '所有字段都不能为空'})
    
    if role not in ['ADMIN', 'MANAGER', 'WORKER']:
        return jsonify({'success': False, 'message': '角色类型无效'})
    
    success, message = db.add_user(username, password, real_name, role)
    return jsonify({'success': success, 'message': message})

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@admin_required
def delete_user(user_id):
    """删除用户"""
    success, message = db.delete_user(user_id)
    return jsonify({'success': success, 'message': message})

@app.route('/api/products', methods=['GET'])
@login_required
def get_products():
    """获取所有产品（支持搜索和分页）"""
    product_type = request.args.get('type', None)
    search = request.args.get('search', None)
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    
    # 限制每页最大数量
    per_page = min(per_page, 100)
    
    products, total = db.get_all_products(
        product_type=product_type,
        search=search,
        page=page,
        per_page=per_page
    )
    
    # 获取库存总和
    total_quantity = db.get_total_quantity(
        product_type=product_type,
        search=search
    )
    
    return jsonify({
        'success': True,
        'data': [{
            'product_id': p[0],
            'name': p[1],
            'product_type': p[2],
            'quantity': p[3],
            'unit': p[4],
            'description': p[5]
        } for p in products],
        'pagination': {
            'page': page,
            'per_page': per_page,
            'total': total,
            'total_pages': (total + per_page - 1) // per_page
        },
        'total_quantity': total_quantity
    })

@app.route('/api/products', methods=['POST'])
@admin_required
def add_product():
    """添加产品"""
    data = request.json
    name = data.get('name', '').strip()
    product_type = data.get('product_type', '').strip()
    unit = data.get('unit', '').strip()
    description = data.get('description', '').strip()
    components = data.get('components', [])
    
    if not name or not unit or not product_type:
        return jsonify({'success': False, 'message': '产品名称、类型和单位不能为空'})
    
    success, message, product_id = db.add_product(name, product_type, unit, description)
    
    if success and product_type == 'FINISHED' and components:
        for comp in components:
            db.add_product_component(product_id, comp['component_id'], comp['quantity'])
    
    return jsonify({'success': success, 'message': message})

@app.route('/api/products/<int:product_id>/components', methods=['GET'])
def get_product_components(product_id):
    """获取成品的配件列表"""
    components = db.get_product_components(product_id)
    return jsonify({
        'success': True,
        'data': [{
            'component_id': c[0],
            'name': c[1],
            'required_quantity': c[2],
            'stock': c[3]
        } for c in components]
    })

@app.route('/api/stock/in', methods=['POST'])
@login_required
def stock_in():
    """入库"""
    data = request.json
    product_name = data.get('product_name', '').strip()
    quantity = data.get('quantity', 0)
    note = data.get('note', '').strip()
    
    # 操作员自动设置为当前登录用户
    operator = session['user']['real_name']
    
    if not product_name:
        return jsonify({'success': False, 'message': '产品名称不能为空'})
    
    try:
        quantity = int(quantity)
        if quantity <= 0:
            return jsonify({'success': False, 'message': '数量必须大于0'})
    except (ValueError, TypeError):
        return jsonify({'success': False, 'message': '数量必须是有效数字'})
    
    success, message = db.stock_in(product_name, quantity, operator, note)
    return jsonify({'success': success, 'message': message})

@app.route('/api/stock/out', methods=['POST'])
@login_required
def stock_out():
    """配件出库（生产）"""
    data = request.json
    product_name = data.get('product_name', '').strip()
    assignments = data.get('assignments', [])
    note = data.get('note', '').strip()
    
    # 操作员自动设置为当前登录用户
    operator = session['user']['real_name']
    
    if not product_name:
        return jsonify({'success': False, 'message': '产品名称不能为空'})
    
    if not assignments or len(assignments) == 0:
        return jsonify({'success': False, 'message': '请至少添加一个生产人员分配'})
    
    # 验证每个分配
    for assignment in assignments:
        if not assignment.get('assigned_to', '').strip():
            return jsonify({'success': False, 'message': '生产人员姓名不能为空'})
        try:
            quantity = int(assignment.get('quantity', 0))
            if quantity <= 0:
                return jsonify({'success': False, 'message': '每个人员的生产数量必须大于0'})
            assignment['quantity'] = quantity
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': '生产数量必须是有效数字'})
    
    success, message = db.stock_out_batch(product_name, assignments, operator, note)
    return jsonify({'success': success, 'message': message})

@app.route('/api/production/complete', methods=['POST'])
def complete_production():
    """完成生产"""
    data = request.json
    trans_id = data.get('trans_id', 0)
    operator = data.get('operator', '').strip()
    
    if not trans_id or not operator:
        return jsonify({'success': False, 'message': '生产ID和操作员不能为空'})
    
    success, message = db.complete_production(trans_id, operator)
    return jsonify({'success': success, 'message': message})

@app.route('/api/defects', methods=['POST'])
@login_required
def record_defects():
    """记录配件损耗"""
    data = request.json
    defects = data.get('defects', [])
    note = data.get('note', '').strip()
    
    # 操作员自动设置为当前登录用户
    operator = session['user']['real_name']
    
    if not defects or len(defects) == 0:
        return jsonify({'success': False, 'message': '请至少添加一个配件损耗'})
    
    # 验证每个损耗记录
    for defect in defects:
        if not defect.get('component_name', '').strip():
            return jsonify({'success': False, 'message': '配件名称不能为空'})
        try:
            quantity = int(defect.get('quantity', 0))
            if quantity <= 0:
                return jsonify({'success': False, 'message': '损耗数量必须大于0'})
            defect['quantity'] = quantity
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': '损耗数量必须是有效数字'})
    
    success, message = db.record_component_defect_batch(defects, operator, note)
    return jsonify({'success': success, 'message': message})

@app.route('/api/transactions', methods=['GET'])
def get_transactions():
    """获取交易历史"""
    trans_type = request.args.get('type', None)
    start_date = request.args.get('start_date', None)
    end_date = request.args.get('end_date', None)
    limit = request.args.get('limit', 50, type=int)
    
    transactions = db.get_transaction_history(
        trans_type=trans_type, 
        start_date=start_date,
        end_date=end_date,
        limit=limit
    )
    return jsonify({
        'success': True,
        'data': [{
            'trans_id': t[0],
            'product_name': t[1],
            'trans_type': t[2],
            'quantity': t[3],
            'operator': t[4],
            'assigned_to': t[5],
            'defect_quantity': t[6],
            'trans_date': t[7],
            'note': t[8]
        } for t in transactions]
    })

@app.route('/api/transactions/export', methods=['GET'])
@login_required
def export_transactions():
    """导出交易历史到Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from flask import send_file
    import io
    from datetime import datetime
    
    trans_type = request.args.get('type', None)
    start_date = request.args.get('start_date', None)
    end_date = request.args.get('end_date', None)
    limit = request.args.get('limit', 1000, type=int)
    
    # 获取交易数据
    transactions = db.get_transaction_history(
        trans_type=trans_type,
        start_date=start_date,
        end_date=end_date,
        limit=limit
    )
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "交易历史"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 30
    
    # 表头样式
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入表头
    headers = ['记录ID', '产品名称', '类型', '数量', '操作员', '分配给', '次品数', '时间', '备注']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 类型映射
    type_map = {
        'IN': '入库',
        'OUT': '配件出库',
        'PRODUCTION': '生产',
        'DEFECT': '次品'
    }
    
    # 写入数据
    for row_idx, trans in enumerate(transactions, start=2):
        ws.cell(row=row_idx, column=1, value=trans[0])  # trans_id
        ws.cell(row=row_idx, column=2, value=trans[1])  # product_name
        ws.cell(row=row_idx, column=3, value=type_map.get(trans[2], trans[2]))  # trans_type
        ws.cell(row=row_idx, column=4, value=trans[3])  # quantity
        ws.cell(row=row_idx, column=5, value=trans[4])  # operator
        ws.cell(row=row_idx, column=6, value=trans[5] if trans[5] else '-')  # assigned_to
        ws.cell(row=row_idx, column=7, value=trans[6] if trans[6] else '-')  # defect_quantity
        ws.cell(row=row_idx, column=8, value=trans[7])  # trans_date
        ws.cell(row=row_idx, column=9, value=trans[8] if trans[8] else '-')  # note
        
        # 数据行居中对齐
        for col in range(1, 10):
            ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center", vertical="center")
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 生成文件名
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    type_suffix = f"_{type_map.get(trans_type, '全部')}" if trans_type else "_全部"
    filename = f"交易历史{type_suffix}_{timestamp}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/api/products/<int:product_id>', methods=['GET'])
def get_product(product_id):
    """获取单个产品信息"""
    product = db.get_product_by_id(product_id)
    if product:
        return jsonify({
            'success': True,
            'data': {
                'product_id': product[0],
                'name': product[1],
                'product_type': product[2],
                'quantity': product[3],
                'unit': product[4],
                'description': product[5]
            }
        })
    return jsonify({'success': False, 'message': '产品不存在'})

@app.route('/api/products/<int:product_id>', methods=['PUT'])
def update_product(product_id):
    """更新产品"""
    data = request.json
    name = data.get('name', '').strip()
    unit = data.get('unit', '').strip()
    description = data.get('description', '').strip()
    quantity = data.get('quantity')
    components = data.get('components', [])
    
    if not name or not unit:
        return jsonify({'success': False, 'message': '产品名称和单位不能为空'})
    
    # 验证数量
    if quantity is not None:
        try:
            quantity = int(quantity)
            if quantity < 0:
                return jsonify({'success': False, 'message': '库存数量不能为负数'})
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': '库存数量必须是有效的整数'})
    
    success, message = db.update_product(product_id, name, unit, description, quantity)
    
    if success and components is not None:
        # 更新成品配件关系
        db.delete_product_components(product_id)
        for comp in components:
            db.add_product_component(product_id, comp['component_id'], comp['quantity'])
    
    return jsonify({'success': success, 'message': message})

@app.route('/api/products/<int:product_id>', methods=['DELETE'])
def delete_product_route(product_id):
    """删除产品"""
    success, message = db.delete_product(product_id)
    return jsonify({'success': success, 'message': message})

@app.route('/api/shipments', methods=['POST'])
@login_required
def create_shipment():
    """成品发货"""
    data = request.json
    shipments = data.get('shipments', [])
    customer_name = data.get('customer_name', '').strip()
    note = data.get('note', '').strip()
    
    # 操作员自动设置为当前登录用户
    operator = session['user']['real_name']
    
    if not customer_name:
        return jsonify({'success': False, 'message': '客户名称不能为空'})
    
    if not shipments or len(shipments) == 0:
        return jsonify({'success': False, 'message': '请至少选择一个成品发货'})
    
    # 验证每个发货项
    for shipment in shipments:
        if not shipment.get('product_name', '').strip():
            return jsonify({'success': False, 'message': '产品名称不能为空'})
        try:
            quantity = int(shipment.get('quantity', 0))
            if quantity <= 0:
                return jsonify({'success': False, 'message': '发货数量必须大于0'})
            shipment['quantity'] = quantity
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': '发货数量必须是有效数字'})
    
    success, message = db.ship_products(shipments, customer_name, operator, note)
    return jsonify({'success': success, 'message': message})

@app.route('/api/shipments', methods=['GET'])
@login_required
def get_shipments():
    """获取发货历史"""
    customer_name = request.args.get('customer_name', None)
    start_date = request.args.get('start_date', None)
    end_date = request.args.get('end_date', None)
    limit = request.args.get('limit', 100, type=int)
    
    shipments = db.get_shipment_history(
        customer_name=customer_name,
        start_date=start_date,
        end_date=end_date,
        limit=limit
    )
    return jsonify({
        'success': True,
        'data': [{
            'trans_id': s[0],
            'product_name': s[1],
            'quantity': s[2],
            'operator': s[3],
            'customer_name': s[4],
            'trans_date': s[5],
            'note': s[6]
        } for s in shipments]
    })

@app.route('/api/shipments/export', methods=['GET'])
@login_required
def export_shipments():
    """导出发货记录到Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from flask import send_file
    import io
    from datetime import datetime
    
    customer_name = request.args.get('customer_name', None)
    start_date = request.args.get('start_date', None)
    end_date = request.args.get('end_date', None)
    limit = request.args.get('limit', 1000, type=int)
    
    # 获取发货数据
    shipments = db.get_shipment_history(
        customer_name=customer_name,
        start_date=start_date,
        end_date=end_date,
        limit=limit
    )
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "发货记录"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 30
    
    # 表头样式
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入表头
    headers = ['记录ID', '产品名称', '数量', '操作员', '客户名称', '发货时间', '备注']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 写入数据
    for row_idx, shipment in enumerate(shipments, start=2):
        ws.cell(row=row_idx, column=1, value=shipment[0])  # trans_id
        ws.cell(row=row_idx, column=2, value=shipment[1])  # product_name
        ws.cell(row=row_idx, column=3, value=shipment[2])  # quantity
        ws.cell(row=row_idx, column=4, value=shipment[3])  # operator
        ws.cell(row=row_idx, column=5, value=shipment[4] if shipment[4] else '-')  # customer_name
        ws.cell(row=row_idx, column=6, value=shipment[5])  # trans_date
        ws.cell(row=row_idx, column=7, value=shipment[6] if shipment[6] else '-')  # note
        
        # 数据行居中对齐
        for col in range(1, 8):
            ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center", vertical="center")
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 生成文件名
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    customer_suffix = f"_{customer_name}" if customer_name else "_全部客户"
    filename = f"发货记录{customer_suffix}_{timestamp}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/api/transactions/<int:trans_id>/cancel', methods=['POST'])
@login_required
def cancel_transaction(trans_id):
    """取消交易记录"""
    # 操作员自动设置为当前登录用户
    operator = session['user']['real_name']
    
    # 只有管理员和仓库管理员可以取消交易
    if session['user']['role'] not in ['ADMIN', 'MANAGER']:
        return jsonify({'success': False, 'message': '权限不足，只有管理员可以取消交易'})
    
    success, message = db.cancel_transaction(trans_id, operator)
    return jsonify({'success': success, 'message': message})

# 客户管理API
@app.route('/api/customers', methods=['GET'])
@login_required
def get_customers():
    """获取所有客户"""
    search = request.args.get('search', None)
    customers = db.get_all_customers(search=search)
    return jsonify({
        'success': True,
        'data': [{
            'customer_id': c[0],
            'name': c[1],
            'contact_person': c[2],
            'phone': c[3],
            'address': c[4],
            'note': c[5],
            'created_at': c[6]
        } for c in customers]
    })

@app.route('/api/customers', methods=['POST'])
@admin_required
def add_customer():
    """添加客户"""
    data = request.json
    name = data.get('name', '').strip()
    contact_person = data.get('contact_person', '').strip()
    phone = data.get('phone', '').strip()
    address = data.get('address', '').strip()
    note = data.get('note', '').strip()
    
    if not name:
        return jsonify({'success': False, 'message': '客户名称不能为空'})
    
    success, message, customer_id = db.add_customer(name, contact_person, phone, address, note)
    return jsonify({'success': success, 'message': message})

@app.route('/api/customers/<int:customer_id>', methods=['GET'])
@login_required
def get_customer(customer_id):
    """获取单个客户信息"""
    customer = db.get_customer_by_id(customer_id)
    if customer:
        return jsonify({
            'success': True,
            'data': {
                'customer_id': customer[0],
                'name': customer[1],
                'contact_person': customer[2],
                'phone': customer[3],
                'address': customer[4],
                'note': customer[5]
            }
        })
    return jsonify({'success': False, 'message': '客户不存在'})

@app.route('/api/customers/<int:customer_id>', methods=['PUT'])
@admin_required
def update_customer(customer_id):
    """更新客户信息"""
    data = request.json
    name = data.get('name', '').strip()
    contact_person = data.get('contact_person', '').strip()
    phone = data.get('phone', '').strip()
    address = data.get('address', '').strip()
    note = data.get('note', '').strip()
    
    if not name:
        return jsonify({'success': False, 'message': '客户名称不能为空'})
    
    success, message = db.update_customer(customer_id, name, contact_person, phone, address, note)
    return jsonify({'success': success, 'message': message})

@app.route('/api/customers/<int:customer_id>', methods=['DELETE'])
@admin_required
def delete_customer(customer_id):
    """删除客户"""
    success, message = db.delete_customer(customer_id)
    return jsonify({'success': success, 'message': message})

@app.route('/api/products/import', methods=['POST'])
@admin_required
def import_products():
    """从Excel导入产品和配件"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '未上传文件'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '未选择文件'})
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': '只支持Excel文件（.xlsx或.xls）'})
    
    try:
        import openpyxl
        import io
        
        # 读取Excel文件
        file_content = file.read()
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        
        # 存储产品ID映射
        product_map = {}
        components_set = set()
        finished_products = {}  # {成品名: [配件列表]}
        
        # 遍历所有工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 扫描所有行，找出所有产品名所在的行
            product_start_rows = []
            for row_idx in range(2, ws.max_row + 1):
                row = list(ws[row_idx])
                if len(row) > 1 and row[1].value:
                    val = str(row[1].value).strip()
                    if val not in ['产品名称', '配件', '']:
                        product_start_rows.append(row_idx)
            
            # 处理每组产品
            for group_idx, start_row in enumerate(product_start_rows):
                # 确定这组产品的结束行
                if group_idx + 1 < len(product_start_rows):
                    end_row = product_start_rows[group_idx + 1] - 1
                else:
                    end_row = ws.max_row
                
                # 从开始行读取所有产品名
                product_columns = {}  # {列索引: 产品名}
                first_row = list(ws[start_row])
                
                col_idx = 1
                while col_idx < len(first_row):
                    cell_value = first_row[col_idx].value
                    if cell_value and str(cell_value).strip() not in ['产品名称', '配件']:
                        product_name = str(cell_value).strip()
                        product_columns[col_idx] = product_name
                        if product_name not in finished_products:
                            finished_products[product_name] = []
                    
                    # 移动到下一组
                    if col_idx + 3 < len(first_row) and first_row[col_idx + 3].value is None:
                        col_idx += 4
                    else:
                        col_idx += 3
                
                # 读取这组产品的所有配件数据
                for row_idx in range(start_row, end_row + 1):
                    row = list(ws[row_idx])
                    
                    # 遍历每个产品列
                    for prod_col_idx, product_name in product_columns.items():
                        # 配件在产品列的下一列
                        component_col_idx = prod_col_idx + 1
                        if component_col_idx < len(row):
                            component_value = row[component_col_idx].value
                            if component_value:
                                component_name = str(component_value).strip()
                                if component_name not in ['产品名称', '配件']:
                                    components_set.add(component_name)
                                    finished_products[product_name].append(component_name)
        
        wb.close()
        
        # 统计信息
        stats = {
            'components_added': 0,
            'components_skipped': 0,
            'finished_added': 0,
            'finished_skipped': 0,
            'relations_added': 0
        }
        
        # 1. 导入所有配件
        for component_name in sorted(components_set):
            success, message, product_id = db.add_product(component_name, 'COMPONENT', '个', '')
            if success:
                product_map[component_name] = product_id
                stats['components_added'] += 1
            else:
                # 如果已存在，获取ID
                product = db.get_product_by_name(component_name)
                if product:
                    product_map[component_name] = product[0]
                stats['components_skipped'] += 1
        
        # 2. 导入所有成品
        for finished_name in sorted(finished_products.keys()):
            success, message, product_id = db.add_product(finished_name, 'FINISHED', '个', '')
            if success:
                product_map[finished_name] = product_id
                stats['finished_added'] += 1
            else:
                product = db.get_product_by_name(finished_name)
                if product:
                    product_map[finished_name] = product[0]
                stats['finished_skipped'] += 1
        
        # 3. 导入配件关系
        for finished_name, components in finished_products.items():
            if finished_name not in product_map:
                continue
            
            finished_id = product_map[finished_name]
            
            for component_name in components:
                if component_name not in product_map:
                    continue
                
                component_id = product_map[component_name]
                
                success = db.add_product_component(finished_id, component_id, 1)
                if success:
                    stats['relations_added'] += 1
        
        return jsonify({
            'success': True,
            'message': '导入成功',
            'stats': stats
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'导入失败: {str(e)}'})

if __name__ == '__main__':
    print("仓库管理系统启动中...")
    print("访问地址: http://localhost:8080")
    app.run(debug=True, host='0.0.0.0', port=8080)

