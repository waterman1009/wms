#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从Excel导入产品配件数据
"""

import openpyxl
import sqlite3
from datetime import datetime

def clear_database(db_name="warehouse.db"):
    """清空所有历史数据（保留用户表）"""
    conn = sqlite3.connect(db_name)
    cursor = conn.cursor()
    
    print("正在清空历史数据...")
    
    # 清空交易记录
    cursor.execute('DELETE FROM transactions')
    print("✓ 清空交易记录")
    
    # 清空配件关系
    cursor.execute('DELETE FROM product_components')
    print("✓ 清空配件关系")
    
    # 清空产品
    cursor.execute('DELETE FROM products')
    print("✓ 清空产品数据")
    
    # 重置自增ID
    cursor.execute('DELETE FROM sqlite_sequence WHERE name IN ("products", "product_components", "transactions")')
    
    conn.commit()
    conn.close()
    print("数据清空完成！\n")

def import_from_excel(excel_path, db_name="warehouse.db"):
    """从Excel导入数据"""
    print(f"正在读取Excel文件: {excel_path}")
    
    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception as e:
        print(f"❌ 无法打开Excel文件: {e}")
        return False
    
    conn = sqlite3.connect(db_name)
    cursor = conn.cursor()
    
    # 存储产品ID映射
    product_map = {}
    components_set = set()
    finished_products = {}  # {成品名: [配件列表]}
    
    print(f"\n发现工作表: {wb.sheetnames}")
    
    # 遍历所有工作表
    for sheet_name in wb.sheetnames:
        print(f"\n处理工作表: {sheet_name}")
        ws = wb[sheet_name]
        
        # 第一步：扫描所有行，找出所有产品名所在的行
        product_start_rows = []  # 存储产品名开始的行号
        for row_idx in range(2, ws.max_row + 1):
            row = list(ws[row_idx])
            # 检查第1列是否有产品名（不是表头）
            if len(row) > 1 and row[1].value:
                val = str(row[1].value).strip()
                if val not in ['产品名称', '配件', '']:
                    product_start_rows.append(row_idx)
        
        print(f"  发现产品组: {len(product_start_rows)}组，行号: {product_start_rows}")
        
        # 第二步：处理每组产品
        for group_idx, start_row in enumerate(product_start_rows):
            # 确定这组产品的结束行（下一组的开始行-1，或工作表结束）
            if group_idx + 1 < len(product_start_rows):
                end_row = product_start_rows[group_idx + 1] - 1
            else:
                end_row = ws.max_row
            
            print(f"  处理第{group_idx+1}组产品: 行{start_row}-{end_row}")
            
            # 从开始行读取所有产品名
            product_columns = {}  # {列索引: 产品名}
            first_row = list(ws[start_row])
            
            col_idx = 1
            while col_idx < len(first_row):
                cell_value = first_row[col_idx].value
                if cell_value and str(cell_value).strip() not in ['产品名称', '配件']:
                    product_name = str(cell_value).strip()
                    product_columns[col_idx] = product_name
                    if product_name not in finished_products:
                        finished_products[product_name] = []
                
                # 移动到下一组（检查分隔符）
                if col_idx + 3 < len(first_row) and first_row[col_idx + 3].value is None:
                    col_idx += 4
                else:
                    col_idx += 3
            
            # 读取这组产品的所有配件数据
            for row_idx in range(start_row, end_row + 1):
                row = list(ws[row_idx])
                
                # 遍历每个产品列
                for prod_col_idx, product_name in product_columns.items():
                    # 配件在产品列的下一列
                    component_col_idx = prod_col_idx + 1
                    if component_col_idx < len(row):
                        component_value = row[component_col_idx].value
                        if component_value:
                            component_name = str(component_value).strip()
                            if component_name not in ['产品名称', '配件']:
                                components_set.add(component_name)
                                finished_products[product_name].append(component_name)
    
    # 1. 先导入所有配件
    print("\n导入配件数据...")
    for component_name in sorted(components_set):
        try:
            cursor.execute(
                'INSERT INTO products (name, product_type, quantity, unit, description) VALUES (?, ?, ?, ?, ?)',
                (component_name, 'COMPONENT', 0, '个', '')
            )
            product_id = cursor.lastrowid
            product_map[component_name] = product_id
            print(f"  ✓ 配件: {component_name} (ID: {product_id})")
        except sqlite3.IntegrityError:
            # 如果已存在，获取ID
            cursor.execute('SELECT product_id FROM products WHERE name = ?', (component_name,))
            result = cursor.fetchone()
            if result:
                product_map[component_name] = result[0]
            print(f"  ⚠ 配件已存在: {component_name}")
    
    # 2. 导入所有成品
    print("\n导入成品数据...")
    for finished_name in sorted(finished_products.keys()):
        try:
            cursor.execute(
                'INSERT INTO products (name, product_type, quantity, unit, description) VALUES (?, ?, ?, ?, ?)',
                (finished_name, 'FINISHED', 0, '个', '')
            )
            product_id = cursor.lastrowid
            product_map[finished_name] = product_id
            print(f"  ✓ 成品: {finished_name} (ID: {product_id})")
        except sqlite3.IntegrityError:
            cursor.execute('SELECT product_id FROM products WHERE name = ?', (finished_name,))
            result = cursor.fetchone()
            if result:
                product_map[finished_name] = result[0]
            print(f"  ⚠ 成品已存在: {finished_name}")
    
    # 3. 导入配件关系
    print("\n导入配件关系...")
    for finished_name, components in finished_products.items():
        if finished_name not in product_map:
            print(f"  ⚠ 成品不存在: {finished_name}")
            continue
        
        finished_id = product_map[finished_name]
        
        for component_name in components:
            if component_name not in product_map:
                print(f"  ⚠ 配件不存在: {component_name}")
                continue
            
            component_id = product_map[component_name]
            
            try:
                cursor.execute(
                    'INSERT INTO product_components (finished_product_id, component_id, quantity) VALUES (?, ?, ?)',
                    (finished_id, component_id, 1)
                )
                print(f"  ✓ {finished_name} <- {component_name} x1")
            except Exception as e:
                print(f"  ❌ 添加关系失败: {e}")
    
    conn.commit()
    conn.close()
    wb.close()
    
    print(f"\n✅ 数据导入完成！")
    print(f"   配件总数: {len(components_set)}")
    print(f"   成品总数: {len(finished_products)}")
    return True

def main():
    """主函数"""
    excel_path = "产品配件.xlsx"
    
    print("=" * 60)
    print("仓库管理系统 - 数据导入工具")
    print("=" * 60)
    
    # 清空数据库
    clear_database()
    
    # 导入Excel数据
    success = import_from_excel(excel_path)
    
    if success:
        print("\n" + "=" * 60)
        print("导入成功！现在可以启动系统了。")
        print("运行命令: python app.py")
        print("=" * 60)
    else:
        print("\n❌ 导入失败，请检查Excel文件格式。")

if __name__ == '__main__':
    main()
